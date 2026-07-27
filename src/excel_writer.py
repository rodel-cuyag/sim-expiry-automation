"""
excel_writer.py
-----------------
Writes report DataFrames into formatted .xlsx workbooks using openpyxl.

Note: values here are pre-computed in Python (not live Excel formulas).
These are operational snapshot reports generated fresh each run from
source data, rather than financial models meant to be edited live in
Excel — so static, correct values are the right choice here.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
import pandas as pd

HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial")

SECTION_HEADER_FILL = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
SECTION_HEADER_FONT = Font(name="Arial", bold=True, size=11)

YELLOW_FILL = PatternFill("solid", start_color="FFFF00", end_color="FFFF00")

# ── "EOD Report" sheet (single dashboard-style summary sheet) ─────

DASH_FONT_NAME = "Work Sans"
DASH_NAVY = "1A2C52"
DASH_LIGHT_BLUE = "E8F0F9"
DASH_GRAY = "555555"
DASH_BLUE_TEXT = "0066CC"
DASH_PURPLE_TEXT = "7B5EA7"
DASH_GREEN_FILL = "00B050"
DASH_RED_FILL = "C00000"

_THIN = Side(style="thin")
_FULL_BORDER = Border(top=_THIN, bottom=_THIN, left=_THIN, right=_THIN)


def _dash_row(label, source=None, highlight=False, yellow=False, no_delta=False, special=None):
    return {
        "label": label, "source": source, "highlight": highlight,
        "yellow": yellow, "no_delta": no_delta, "special": special,
    }


# One entry per sheet row: display label, the matching eod_df metric label
# to pull the "Today" value from (written as a literal value, not a
# formula), and any special styling. "Failed" and "Total Completed Calls"
# are included even though the reference dashboard doesn't show them, for
# consistency with the metrics computed elsewhere (see plan doc).
DASHBOARD_ROWS = [
    _dash_row("Calls Dialled - Target", "Calls Dialed - Target", no_delta=True),
    _dash_row("Calls Dialled - Actual", "Calls Dialed - Actual"),
    _dash_row("Calls Connected", "Calls Connected", highlight=True),
    _dash_row("No Answer (all retries exhausted)", "No Answer"),
    _dash_row("Busy", "Busy"),
    _dash_row("Failed", "Failed"),
    _dash_row("System Errors", special="system_errors"),
    _dash_row("Total Completed Calls", "Total Completed Calls"),
    _dash_row("Total Call Duration (minutes)", "Total Call Duration (minutes)"),
    _dash_row("Avg. Call Duration - Connected (seconds)", "Avg. Call Duration - Connected (seconds)"),
    _dash_row("Connection Rate (Connected / Dialled)", "Connection Rate (Connected / Dialed)", highlight=True),
    _dash_row("Agreed to Keep SIM Active (count)", "Agreed to Keep SIM Active (count)", highlight=True),
    _dash_row("Conversion Rate (Agreed / Connected)", "Conversion Rate (Agreed / Connected)", highlight=True),
    _dash_row("Retries Queued for Tomorrow", "Retries Queued for Tomorrow", no_delta=True),
    _dash_row("__SECTION__", special="FINOPS"),
    _dash_row("LLM Inference Cost", "LLM Inference Cost (USD)", yellow=True),
    _dash_row("Total Daily Spend", "Total Daily Spend (USD)", yellow=True, highlight=True),
    _dash_row("__SECTION__", special="ISSUES & CHANGES"),
    _dash_row("Open P0 issues", "Open P0 Issues", yellow=True),
    _dash_row("Open P1 issues", "Open P1 Issues", yellow=True),
    _dash_row("Changes deployed today", "Changes Deployed Today", yellow=True, no_delta=True),
    _dash_row("Changes pending approval for tomorrow", "Changes Pending Approval for Tomorrow", yellow=True, no_delta=True),
    _dash_row("__SECTION__", special="TOMORROW'S PLAN"),
    _dash_row("Target call volume", "Target Call Volume", no_delta=True),
    _dash_row("Expected list from Globe (ETA)", "Expected List from Globe (ETA)", no_delta=True),
    _dash_row("Calling window", "Calling Window", no_delta=True),
    _dash_row("Phase gate status", "Phase Gate Status", no_delta=True, special="phase_gate"),
]

# Call Detail Log column widths matching the reference format exactly
# (unspecified columns, e.g. G/J, keep their auto-computed width).
CALL_DETAIL_LOG_COLUMN_WIDTHS = {
    "A": 40, "B": 18, "C": 13, "D": 23, "E": 29, "F": 24, "H": 37, "I": 19,
}


def resolve_output_path(path: Path) -> Path:
    """
    If *path* already exists, append *(1)*, *(2)*, etc. before the
    extension so the existing file is never silently overwritten.
    """
    if not path.exists():
        return path
    stem = path.stem
    suffix = path.suffix
    parent = path.parent
    n = 1
    while True:
        candidate = parent / f"{stem} ({n}){suffix}"
        if not candidate.exists():
            return candidate
        n += 1


def _write_dataframe(ws, df: pd.DataFrame):
    """Writes a DataFrame to a worksheet with a styled header row and auto-width columns."""
    ws.append(list(df.columns))
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    for row in df.itertuples(index=False):
        ws.append(list(row))

    for i, col in enumerate(df.columns, start=1):
        if not df.empty:
            col_len = df[col].astype(str).str.len().max()
            col_len = 0 if pd.isna(col_len) else col_len
        else:
            col_len = 0
        max_len = max(col_len, len(str(col)))
        ws.column_dimensions[get_column_letter(i)].width = min(max_len + 4, 45)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BODY_FONT

    ws.freeze_panes = "A2"


def _write_eod_summary_sheet(ws, eod_df: pd.DataFrame):
    """
    Writes the "EOD Report" sheet: big navy header, Today/Yesterday/Delta
    comparison table. "Today" values are literal values computed in Python
    (from eod_df), not formulas — there's no separate source sheet to pull
    from. Yesterday/Delta and the "Day X of N" campaign counter are left as
    blank/bracket placeholders — the pipeline doesn't track prior-day report
    data. "System Errors" is the one live formula, referencing the Target/
    Actual rows on this same sheet.
    """
    value_of = dict(eod_df.itertuples(index=False))

    ws.merge_cells("A1:D1")
    title = ws["A1"]
    title.value = "End-of-Day Report"
    title.font = Font(name=DASH_FONT_NAME, size=36, color=DASH_NAVY)
    title.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 42

    subtitle = ws["A2"]
    # 14 is the fixed known campaign length; the current day number isn't
    # tracked anywhere yet, so it stays a fill-in-by-hand placeholder.
    subtitle.value = f"Date: {value_of['Report Period']}    Day: [N] of 14"
    subtitle.font = Font(name=DASH_FONT_NAME, size=18, bold=True, color=DASH_GRAY)

    navy_fill = PatternFill("solid", start_color=DASH_NAVY, end_color=DASH_NAVY)
    header_font = Font(name=DASH_FONT_NAME, size=18, bold=True, color="FFFFFF")
    for col, text in enumerate(["Metric", "Today", "Yesterday", "Δ"], start=1):
        cell = ws.cell(row=3, column=col, value=text)
        cell.font = header_font
        cell.fill = navy_fill
    ws.cell(row=3, column=1).border = Border(top=_THIN, left=_THIN)
    ws.cell(row=3, column=2).border = Border(top=_THIN)
    ws.cell(row=3, column=3).border = Border(top=_THIN)
    ws.cell(row=3, column=4).border = Border(top=_THIN, right=_THIN)

    subheader_font = Font(name=DASH_FONT_NAME, size=18, color="F5F5F5")
    for col, text in enumerate(["", "T0", "T-1", "vs T-1"], start=1):
        cell = ws.cell(row=4, column=col, value=text)
        cell.font = subheader_font
        cell.fill = navy_fill
    ws.cell(row=4, column=1).border = Border(left=_THIN)
    ws.cell(row=4, column=4).border = Border(right=_THIN)

    def _fill(color):
        return PatternFill("solid", start_color=color, end_color=color)

    row_of = {}
    body_row = 5
    for spec in DASHBOARD_ROWS:
        if spec["label"] == "__SECTION__":
            ws.merge_cells(start_row=body_row, start_column=2, end_row=body_row, end_column=4)
            a = ws.cell(row=body_row, column=1)
            a.fill = navy_fill
            a.border = Border(left=_THIN)
            b = ws.cell(row=body_row, column=2, value=spec["special"])
            b.font = header_font
            for col in (2, 3, 4):
                ws.cell(row=body_row, column=col).fill = navy_fill
            ws.cell(row=body_row, column=4).border = Border(right=_THIN)
            body_row += 1
            continue

        if spec["source"]:
            row_of[spec["source"]] = body_row

        highlight = spec["highlight"]
        band_color = DASH_LIGHT_BLUE if highlight else "FFFFFF"
        label_font = Font(name=DASH_FONT_NAME, size=18, color=DASH_GRAY)
        value_font = Font(
            name=DASH_FONT_NAME, size=18, bold=highlight,
            color=DASH_BLUE_TEXT if highlight else DASH_NAVY,
        )
        delta_font = Font(name=DASH_FONT_NAME, size=18, color=DASH_PURPLE_TEXT)

        a = ws.cell(row=body_row, column=1, value=spec["label"])
        a.font = label_font
        a.fill = _fill(band_color)
        a.border = _FULL_BORDER

        if spec["special"] == "system_errors":
            value = f"=MAX(0,B{row_of['Calls Dialed - Target']}-B{row_of['Calls Dialed - Actual']})"
        else:
            value = value_of[spec["source"]]
        b = ws.cell(row=body_row, column=2, value=value)
        b.font = value_font
        b.fill = YELLOW_FILL if spec["yellow"] else _fill(band_color)
        b.alignment = Alignment(horizontal="left")
        b.border = _FULL_BORDER

        c = ws.cell(row=body_row, column=3, value=None)
        c.font = delta_font
        c.fill = _fill(band_color)
        c.border = _FULL_BORDER

        d = ws.cell(row=body_row, column=4, value="—" if spec["no_delta"] else "[+/-N]")
        d.font = delta_font
        d.fill = _fill(band_color)
        d.border = _FULL_BORDER

        if spec["special"] == "phase_gate":
            ws.conditional_formatting.add(
                f"B{body_row}",
                FormulaRule(
                    formula=[f'ISNUMBER(SEARCH("GO",B{body_row}))'],
                    fill=_fill(DASH_GREEN_FILL),
                    font=Font(name=DASH_FONT_NAME, bold=True, color="FFFFFF"),
                ),
            )
            ws.conditional_formatting.add(
                f"B{body_row}",
                FormulaRule(
                    formula=[f'ISNUMBER(SEARCH("Hold",B{body_row}))'],
                    fill=_fill(DASH_RED_FILL),
                    font=Font(name=DASH_FONT_NAME, bold=True, color="FFFFFF"),
                ),
            )

        body_row += 1

    ws.column_dimensions["A"].width = 72
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 18.5
    ws.column_dimensions["D"].width = 12


def write_eod_report_sheets(eod_df: pd.DataFrame, call_detail_df: pd.DataFrame, output_path):
    """
    Creates the EOD-mode 2-sheet workbook: 'EOD Report' (dashboard-style
    summary sheet) and 'Call Detail Log'.
    """
    wb = Workbook()

    eod_sheet = wb.active
    eod_sheet.title = "EOD Report"
    _write_eod_summary_sheet(eod_sheet, eod_df)

    detail_sheet = wb.create_sheet("Call Detail Log")
    _write_dataframe(detail_sheet, call_detail_df)
    _apply_date_format(detail_sheet, list(call_detail_df.columns), ["Call Date (PHT)"])
    for col, width in CALL_DETAIL_LOG_COLUMN_WIDTHS.items():
        detail_sheet.column_dimensions[col].width = width

    wb.save(output_path)
    return output_path


def write_priority_list_sheet(df: pd.DataFrame, output_path, sheet_name: str, date_columns=None):
    """
    Creates a single-sheet workbook for the valid Priority List.
    date_columns: optional list of column names to format as plain
    dates (YYYY-MM-DD) instead of openpyxl's default datetime display.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    _write_dataframe(ws, df)
    _apply_date_format(ws, list(df.columns), date_columns)
    wb.save(output_path)
    return output_path


_SHEET_TITLES = {
    "summary": "Summary",
    "invalid": "Invalid Data",
    "expired": "Expired Numbers",
    # EOD validation sheet titles
    "join_summary": "Join Summary",
    "field_completeness": "Field Completeness",
    "calculation_audit": "Calculation Audit",
    "data_quality_issues": "Data Quality Issues",
}


def write_validation_report(sheets: dict, output_path, date_columns=None):
    """
    Creates a multi-sheet validation report workbook.
    Each key in *sheets* is a sheet identifier; the value is a DataFrame.
    The first key determines the first (active) sheet; subsequent keys become
    additional sheets in iteration order.
    """
    keys = list(sheets.keys())
    if not keys:
        raise ValueError("At least one sheet is required")

    wb = Workbook()
    first_key = keys[0]
    wb.active.title = _SHEET_TITLES.get(first_key, first_key.replace("_", " ").title())
    _write_dataframe(wb.active, sheets[first_key])
    if date_columns:
        _apply_date_format(wb.active, list(sheets[first_key].columns), date_columns)

    for key in keys[1:]:
        title = _SHEET_TITLES.get(key, key.replace("_", " ").title())
        ws = wb.create_sheet(title)
        _write_dataframe(ws, sheets[key])
        if date_columns:
            _apply_date_format(ws, list(sheets[key].columns), date_columns)

    wb.save(output_path)
    return output_path


# ── CSV writers (Priority List) ───────────────────────────────────


def write_priority_list_csv(df: pd.DataFrame, output_path):
    """Write the Priority List (valid records) to a CSV file."""
    df.to_csv(output_path, index=False)
    return output_path


# ── Excel helpers ─────────────────────────────────────────────────


def _apply_date_format(ws, column_names, date_columns):
    """Apply YYYY-MM-DD formatting to cells in date_columns (skip header row)."""
    if not date_columns:
        return
    for col_name in date_columns:
        if col_name not in column_names:
            continue
        col_letter = get_column_letter(column_names.index(col_name) + 1)
        for cell in ws[col_letter][1:]:
            cell.number_format = "YYYY-MM-DD"