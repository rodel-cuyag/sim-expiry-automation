"""
prior_day.py
-------------
Best-effort lookup of the previous day's already-generated EOD Report
workbook, so the current run's "Yesterday" column can be filled in
without recomputing anything from the source data.

Two of the dashboard rows ("System Errors", "Retries Queued for
Tomorrow") are written as live Excel formulas rather than literal
values (see excel_writer.py) — openpyxl can't evaluate a formula's
result, only Excel/LibreOffice can, on open. Those two rows are simply
skipped here and left for the caller to treat as "no data available",
same as a missing prior report.
"""

from openpyxl import load_workbook

from src import config


def _find_previous_report(agent_id, previous_date):
    """
    Returns the most recently modified EOD report file for *previous_date*
    (searching recursively across that date's run-time subfolders, and
    covering any `resolve_output_path` rerun suffixes like " (1).xlsx"
    within a given run), or None if no prior report exists.
    """
    eod_dir = config.get_eod_output_dir(previous_date, previous_date)
    if not eod_dir.exists():
        return None

    stem = config.OUTPUT_FILENAME_TEMPLATE_SINGLE.format(
        agent_id=agent_id, start_date=previous_date,
    ).rsplit(".xlsx", 1)[0]
    candidates = list(eod_dir.rglob(f"{stem}*.xlsx"))
    if not candidates:
        return None
    return max(candidates, key=lambda p: p.stat().st_mtime)


def load_previous_day_values(agent_id, previous_date) -> dict:
    """
    Returns {metric_label: value} read from Column A/B of the previous
    day's "EOD Report" sheet, for every row with a literal (non-formula)
    value in Column B. Returns {} if no prior report is found or it
    can't be read.
    """
    path = _find_previous_report(agent_id, previous_date)
    if path is None:
        return {}

    try:
        wb = load_workbook(path, data_only=False)
        if "EOD Report" not in wb.sheetnames:
            return {}
        ws = wb["EOD Report"]

        values = {}
        for row in ws.iter_rows(min_row=5, max_col=2):
            label_cell, value_cell = row[0], row[1]
            if label_cell.value is None:
                continue  # blank / section-header row
            value = value_cell.value
            if isinstance(value, str) and value.startswith("="):
                continue  # live formula; can't reliably read its result
            values[label_cell.value] = value
        return values
    except Exception:
        return {}
