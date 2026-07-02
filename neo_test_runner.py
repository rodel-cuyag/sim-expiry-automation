"""
Neo Agentic Workflow – Test Case Runner
========================================
Usage:
    python neo_test_runner.py

Requirements:
    pip install requests openpyxl

What it does:
    1. Reads an Excel file you specify (INPUT_FILE).
    2. Iterates every sheet that matches TARGET_SHEETS.
    3. For each row, takes the value in INPUT_QUERY_COL and sends it to
       the Neo agentic-flow API as payload.input_value → post.query.
    4. Parses the API response and writes:
         - Neo Topic Classification  ← transactional_type or conversational_type
         - Neo Output Response        ← post_response
    5. Saves the updated workbook to OUTPUT_FILE.
"""

import copy
import json
import time
import requests
from openpyxl import load_workbook

# ─────────────────────────────────────────
#  CONFIGURATION — edit these before running
# ─────────────────────────────────────────
INPUT_FILE  = "Test Inputs-Outputs.xlsx"   # path to your uploaded Excel
OUTPUT_FILE = "Test_Outputs_Neo.xlsx"       # where results are saved

API_URL = (
    "https://agentic-flow.airyn-studio.dev.neo.build"
    "/api/v1/run/1738d6c9-2e9f-4e2c-97df-8dabe3b64eed"
)
API_KEY = "38c230e5-85c9-4c32-a77d-0826cb6ff9f7"   # ← replace with your real key

# Only process sheets whose names contain one of these strings (case-insensitive).
# Leave empty  []  to process ALL sheets.
TARGET_SHEETS = ["PLDT_OUTLOOK", "Dify vs Neo"]

# Column header names (must match the Excel headers exactly, case-insensitive search)
INPUT_QUERY_COL         = "Input Query"
OUTPUT_CLASSIFICATION   = "Neo Topic Classification"
OUTPUT_RESPONSE         = "Neo Output Response"

# Seconds to wait between API calls (be polite to the server)
SLEEP_BETWEEN_CALLS = 1.0

# ─────────────────────────────────────────
#  BASE PAYLOAD — the full JSON skeleton
#  Only post.query is replaced per row.
# ─────────────────────────────────────────
BASE_INPUT_VALUE = {
    "user_metadata": {
        "username": "rodel.cuyag@onebyzero.ai",
        "name": "Rodel Cuyag",
        "handle": "@rodel.cuyag@onebyzero.ai",
        "user_id": "rodel.cuyag@onebyzero.ai"
    },
    "customer_metadata": {
        "description": "pldt x",
        "name": "pldt"
    },
    "products": [],
    "product_categories": [
        {"name": "Fixed Postpaid Plans", "description": None},
        {"name": "Value-added Services", "description": None},
        {"name": "Fixed Postpaid Plans", "description": None},
        {"name": "Devices", "description": None}
    ],
    "product_subcategories": [],
    "customer_intents": [
        {
            "name": "Connectivity",
            "description": (
                "Related to network issues, signal strength, internet speed, etc. "
                "Keywords: 1Bar, 2Bar, 3Bar, 3G, 4Bar, 4G, 4G LTE, 5G, afk, ..."
            )
        },
        {
            "name": "FollowUp",
            "description": (
                "Inquiries about status updates, waiting times, or repeated requests. "
                "Keywords: almost months, almost mos, ang tagal, anong petsa na, ..."
            )
        },
        {
            "name": "Billing Aftersales",
            "description": (
                "Questions or issues related to payments, accounts, or after-sales services. "
                "Keywords: accnt, account, activate, activation, adjustment, ..."
            )
        }
    ],
    "off_topics": [],
    "channel_metadata": {
        "type": "outlook",
        "username": "demo01@onebyzero.ai",
        "handle": "@demo01@onebyzero.ai",
        "post_response_character_limit": 1960,
        "post_response_character_limit_description": None,
        "id": "160"
    },
    "post": {
        "query": "",          # ← filled in per row
        "id": "5372",
        "previous_response": None,
        "previous_response_confidence_score": None,
        "post_features": None
    },
    "rephrased_remark": {
        "persona": "default",
        "rephrase_context": (
            "Professional, efficient communicator. Clear, accurate, prompt responses. "
            "Fact-based, concise answers. Consistent tone, no emojis. Escalate complex "
            "issues. Avoid opinions. Goal: Provide efficient, helpful service representing "
            "the company well."
        ),
        "feedback": None
    },
    "case_creation_metadata": "",
    "conversation_history": [],
    "current_email_attachments": [],
    "summary": ""
}


# ─────────────────────────────────────────
#  HELPERS
# ─────────────────────────────────────────

def find_header_row(sheet):
    """
    Scan the first 10 rows to find a row that contains 'Input Query'.
    Returns (row_index_1based, col_map) where col_map is {header: col_index}.
    """
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True), start=1):
        headers = [str(c).strip() if c is not None else "" for c in row]
        if any(INPUT_QUERY_COL.lower() in h.lower() for h in headers):
            col_map = {h: i + 1 for i, h in enumerate(headers) if h}
            return row_idx, col_map
    return None, {}


def col_index(col_map, name):
    """Case-insensitive column lookup."""
    for key, idx in col_map.items():
        if key.lower() == name.lower():
            return idx
    return None


def call_api(query: str) -> dict:
    """
    Build the payload with the given query and POST to the Neo API.
    Returns the parsed JSON response dict, or {} on error.
    """
    input_value = copy.deepcopy(BASE_INPUT_VALUE)
    input_value["post"]["query"] = query

    payload = {
        "input_value": json.dumps(input_value),
        "output_type": "chat",
        "input_type": "chat"
    }
    headers = {
        "Content-Type": "application/json",
        "api-key": API_KEY
    }

    try:
        resp = requests.post(API_URL, json=payload, headers=headers, timeout=60)
        resp.raise_for_status()
        return resp.json()
    except requests.exceptions.RequestException as exc:
        print(f"    ✗ API error: {exc}")
        return {}


def parse_response(raw: dict):
    """
    Extract Neo Topic Classification and Neo Output Response from the raw API JSON.

    Actual response structure (confirmed via Postman):
        raw
        └── outputs[0]
            └── outputs[0]
                └── results
                    └── message
                        └── text   ← stringified JSON string

    The text string, once parsed, looks like:
        {
            "post_response": "...",
            "transactional_type": "...",
            "conversational_type": "...",
            ...
        }

    classification = transactional_type (if non-empty) else conversational_type
    response       = post_response
    """
    result_obj = None

    # Primary path: outputs[0]["outputs"][0]["results"]["message"]["text"]
    try:
        text = raw["outputs"][0]["outputs"][0]["results"]["message"]["text"]
        if isinstance(text, str):
            result_obj = json.loads(text)
        elif isinstance(text, dict):
            result_obj = text
    except (KeyError, IndexError, TypeError, json.JSONDecodeError):
        pass

    # Fallback: outputs[0]["outputs"][0]["outputs"]["message"]["message"]
    if result_obj is None:
        try:
            msg = raw["outputs"][0]["outputs"][0]["outputs"]["message"]["message"]
            if isinstance(msg, str):
                result_obj = json.loads(msg)
            elif isinstance(msg, dict):
                result_obj = msg
        except (KeyError, IndexError, TypeError, json.JSONDecodeError):
            pass

    # Fallback: artifacts["message"] at outputs[0]["outputs"][0]
    if result_obj is None:
        try:
            msg = raw["outputs"][0]["outputs"][0]["artifacts"]["message"]
            if isinstance(msg, str):
                result_obj = json.loads(msg)
            elif isinstance(msg, dict):
                result_obj = msg
        except (KeyError, IndexError, TypeError, json.JSONDecodeError):
            pass

    if not result_obj:
        print(f"    ✗ Could not parse result. Raw keys: {list(raw.keys())}")
        return "PARSE ERROR", "PARSE ERROR"

    # Extract fields — format as "Topic Name (Transactional)" or "Topic Name (Conversational)"
    transactional  = (result_obj.get("transactional_type")  or "").strip()
    conversational = (result_obj.get("conversational_type") or "").strip()
    response       =  result_obj.get("post_response", "")

    if transactional:
        classification = f"{transactional} (Transactional)"
    elif conversational:
        classification = f"{conversational} (Conversational)"
    else:
        classification = ""

    return classification, str(response).strip()


def should_process_sheet(sheet_name: str) -> bool:
    if not TARGET_SHEETS:
        return True
    return any(t.lower() in sheet_name.lower() for t in TARGET_SHEETS)


# ─────────────────────────────────────────
#  MAIN
# ─────────────────────────────────────────

def main():
    print(f"Loading workbook: {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)

    for sheet_name in wb.sheetnames:
        if not should_process_sheet(sheet_name):
            print(f"\nSkipping sheet: {sheet_name}")
            continue

        ws = wb[sheet_name]
        print(f"\n{'='*60}")
        print(f"Sheet: {sheet_name}")

        header_row, col_map = find_header_row(ws)
        if header_row is None:
            print("  ⚠ No header row with 'Input Query' found — skipping.")
            continue

        # Locate required columns
        q_col  = col_index(col_map, INPUT_QUERY_COL)
        nc_col = col_index(col_map, OUTPUT_CLASSIFICATION)
        nr_col = col_index(col_map, OUTPUT_RESPONSE)

        if q_col is None:
            print(f"  ⚠ Column '{INPUT_QUERY_COL}' not found — skipping.")
            continue

        # If output columns don't exist yet, append them
        if nc_col is None:
            nc_col = ws.max_column + 1
            ws.cell(row=header_row, column=nc_col, value=OUTPUT_CLASSIFICATION)
            col_map[OUTPUT_CLASSIFICATION] = nc_col
            print(f"  + Added column '{OUTPUT_CLASSIFICATION}' at col {nc_col}")

        if nr_col is None:
            nr_col = ws.max_column + 1
            ws.cell(row=header_row, column=nr_col, value=OUTPUT_RESPONSE)
            col_map[OUTPUT_RESPONSE] = nr_col
            print(f"  + Added column '{OUTPUT_RESPONSE}' at col {nr_col}")

        # Iterate data rows
        data_rows = list(ws.iter_rows(min_row=header_row + 1, values_only=False))
        processed = 0

        for row in data_rows:
            query_cell = row[q_col - 1]
            query = query_cell.value

            # Skip blank or non-string cells
            if not query or not str(query).strip():
                continue

            query = str(query).strip()
            print(f"\n  Row {query_cell.row}: {query[:80]}{'…' if len(query) > 80 else ''}")

            # Call the API
            raw = call_api(query)

            if raw:
                classification, response = parse_response(raw)
                print(f"    Classification : {classification}")
                print(f"    Response       : {response[:100]}{'…' if len(response) > 100 else ''}")
            else:
                classification, response = "API ERROR", "API ERROR"

            # Write results back into the sheet
            ws.cell(row=query_cell.row, column=nc_col, value=classification)
            ws.cell(row=query_cell.row, column=nr_col, value=response)

            processed += 1
            time.sleep(SLEEP_BETWEEN_CALLS)

        print(f"\n  ✓ Processed {processed} rows in '{sheet_name}'")

    wb.save(OUTPUT_FILE)
    print(f"\n{'='*60}")
    print(f"Done! Results saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()